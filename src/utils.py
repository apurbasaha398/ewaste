import pandas as pd
import googlemaps
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from src.exception import CustomException
import sys

API_KEY = "Your API Key"

def calculate_transportation_cost(start, end):
    # Initialize the Google Maps client with your API key
    gmaps = googlemaps.Client(key=API_KEY)

    # Prepare a DataFrame to store the distances
    cost_matrix = pd.DataFrame(index=start.keys(), columns=end.keys())
    cost_per_mile_per_ton = 0.36

    # Calculate distances in miles
    for start_loc, start_coords in start.items():
        for end_loc, end_coords in end.items():
            result = gmaps.distance_matrix(origins=[start_coords], destinations=[end_coords], mode='driving', units='imperial')
            # Check if the response contains valid data
            if result['rows'][0]['elements'][0]['status'] == 'OK':
                distance = result['rows'][0]['elements'][0]['distance']['value'] / 1609.34  # Convert meters to miles
                cost_matrix.at[start_loc, end_loc] = distance * cost_per_mile_per_ton
            else:
                # Handle cases where distance is not available
                print(start_loc, end_loc)
                cost_matrix.at[start_loc, end_loc] = None  # or assign a default value

    # return the cost matrix
    return cost_matrix.reset_index().rename(columns={cost_matrix.index.name:'code'})

def add_sheet_to_excelbook(file_path, sheet_name, new_data):
    """
    Adds or rewrites a sheet in an existing Excel workbook with new data.

    :param file_path: Path to the Excel workbook file.
    :param sheet_name: Name of the sheet to be added or rewritten.
    :param new_data: Data to be written into the sheet (as a DataFrame).
    """
    try:
        wb = load_workbook(file_path)  # Open workbook
        if sheet_name in wb.sheetnames:  # If sheet exists, delete it
            del wb[sheet_name]
            #print(f"Sheet '{sheet_name}' successfully removed in '{file_path}'.")

        ws = wb.create_sheet(title=sheet_name)  # Create new sheet

        rows = dataframe_to_rows(new_data, index=False, header=True) # Write dataframe as rows

        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value) #Add... the 2, 4 are the offset, similar to the startrow and startcol in your code
                    
        wb.save(file_path)
        #print(f"Sheet '{sheet_name}' successfully rewritten in '{file_path}'.")

    except Exception as e:
            raise CustomException(e, sys)
    
